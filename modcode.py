'''this script is used to to create the .xlsx file with the raw data imported,
 get the path and download all the files and to check for any corrupted files.
This file contains 2 main functions- get_path(), download_file()'''


import openpyxl

import os
import glob
import requests 
def get_path():
    '''
    Gets the path and prints the file list which is the excel file name
    Variables:
    file_list: path as the excel file name(data type:str )
    path: directory location(data type: str)
    '''
    global file_list
    global path
    path="/home/unknown/Desktop/iom_files/"
    file_list = glob.glob(path+"*.xlsx")
    print(file_list)



    
def download_file():
    '''
    To create the directory with the folder name only if it's unique,
    to change current directory to the new path directory
    to chack for values in column 8, else check column 9
    Variables:
    folder_name:splitting the entire path to get the unique code(data type: str)
    wrkbk: spreadsheet with file names
    col_names: datatype(list)
    File_path:ftp link
    '''
    for f in file_list:
        folder_name=f.split("/")[-1].split(".")[0]
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
    wrkbk = openpyxl.load_workbook(f)
    os.chdir(path+folder_name)
    sheet = wrkbk.active
    col_names = []
    print(sheet.max_row)
    for i in range (2,sheet.max_row+1):
        file_path=sheet.cell(row=i,column=8).value
        print(file_path)
        if file_path == None:
            file_path=sheet.cell(row=i,column=9).value
        print(file_path)
        files=file_path.split(";")
        print(files)

        for ftp in files:
            try:
                os.system("aria2c -c --file-allocation=none -x 16 http://"+ftp)
            except:
                print(ftp + "Didnt download the files")
        wrkbk.save(f)

def getMd5(file_path):
    m = hashlib.md5()
    with open(file_path,'rb') as f:
        lines = f.read()
        m.update(lines)
    md5code = m.hexdigest()
    if md5code!=ena_md5:
        print(f)
    else:
        print("The files look okay"+str(f))





def corruption():
    '''to check for any corrupted files'''
    path =os.getcwd()
    print(path)

    def getMd5(file_path):
        m = hashlib.md5()
        with open(file_path,'rb') as f:
            lines = f.read()
            m.update(lines)
        md5code = m.hexdigest()
        if md5code!=ena_md5:
            print(f)
        else:
            print("The files look okay"+str(f))

    file_list = glob.glob("*.fastq.gz")
    print(file_list)
    for f in file_list:
        file_ids=f.split(".fastq.")[0]
        file_id=file_ids.split("_")
        print(file_id)
        part=file_id[1]
        x=requests.get('https://www.ebi.ac.uk/ena/portal/api/filereport?accession='+file_id[0]+'&result=read_run&fields=fastq_md5')
        print(x)
        ena_md5=x.text.splitlines()[1].split('\t')[1].split(";")[int(part)-1]
        file_path=(path+"/"+f)
        print("hello")
        getMd5(file_path)


#main directory
get_path()
download_file()
corruption()
