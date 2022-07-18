import openpyxl

import os
import glob
path="/home/ubuntu/iom/"
file_list = glob.glob(path+"*.xlsx")
print(file_list)
for f in file_list:
    folder_name=f.split("/")[-1].split(".")[0]
    os.mkdir(path+folder_name)
    wrkbk = openpyxl.load_workbook(f)
    os.chdir(path+folder_name)
    sheet = wrkbk.active
    col_names = []
    print(sheet.max_row)
    for i in range (2,sheet.max_row+1):
        file_path=sheet.cell(row=i,column=7).value
        print(file_path)
        if file_path == None:
            file_path=sheet.cell(row=i,column=8).value

        print(file_path)
        files=file_path.split(";")
        print(files)
        for ftp in files:
            try:
                os.system("aria2c -c --file-allocation=none -x 16 http://"+ftp)
            except:
                print(ftp + "Didnt download the files") 
    wrkbk.save(f)
